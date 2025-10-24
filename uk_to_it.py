# -*- coding: utf-8 -*-
import os, re, sys, csv, random
import numpy as np
import pandas as pd
from collections import defaultdict
from openpyxl import load_workbook

# ==================== BUSINESS RULES ====================

# default item tail (can be overridden from Find_and_Replace.xlsx -> sheet "settings", key=item_tail_it)
ITEM_TAIL = " - Rear Sides & Rear Window - Custom Fit, UV Protection, Heat & Glare Reduction, High Performance"

# default generic keywords (can be overridden from settings: generic_keywords_default)
GENERIC_KEYWORDS_DEFAULT = "pellicola oscurante vetri auto, pellicola vetri auto, oscuramento vetri auto, oscurare vetri auto"

def normalize_item_name(name: str, item_tail: str = "") -> str:
    if not isinstance(name, str) or "TINTCOM" not in name:
        return name
    m = re.search(r"for\s+(.+?)\s+-\s+(?:\((\d+%[^)]*)\)\s+-\s+)?", name, flags=re.I)
    if not m:
        return name
    car = m.group(1).strip()
    shade = (m.group(2) or "").strip()
    tail = item_tail or ITEM_TAIL
    return (f"TINTCOM Pre-Cut Window Tint Film for {car} - ({shade}){tail}"
            if shade else f"TINTCOM Pre-Cut Window Tint Film for {car}{tail}")

# Ако искаш италиански етикети – смени стойностите вдясно.
# NOTE: Tone map moved to Find_and_Replace.xlsx -> sheet "tone_map"
# TONE_MAP removed from code.

# IT_KEYWORD_SETS / DEFAULT_GENERIC_KEYWORDS moved into settings (generic_keywords_default)
# ...existing code...

COLUMNS_TO_COPY = []  # ако е празно -> копира всички съвпадащи по име
PRICE_COLS_HINTS = ["price","our_price","standard_price","list_price",
                    "minimum_seller_allowed_price","maximum_seller_allowed_price"]

# ==================== ЛОГ/КОНСТАНТИ ====================

APPLIED_LOG = defaultdict(int)

# колони само с ТОЧНА замяна (никакъв regex)
EXACT_ONLY_FIELDS = {
    "variation_theme","package_level","material_type","color_map","unit_count_type",
    "length_longer_edge_unit_of_measure","width_shorter_edge_unit_of_measure",
    "package_length_unit_of_measure","fulfillment_center_id","is_fragile",
    "package_weight_unit_of_measure","country_of_origin",
    "compliance_media_content_type1","gpsr_safety_attestation",
    "gpsr_manufacturer_reference_email_address",
    "compliance_media_source_location1","compliance_media_content_language1",
    "product_tax_code","condition_type"
}

# дълги текстови колони – позволяваме regex
LONG_TEXT_FIELDS = {
    "item_name","product_description","bullet_point1","bullet_point2",
        "bullet_point3","bullet_point4","bullet_point5","generic_keywords",
    "generic_keyword","color_name"
}

GENERIC_KEYWORD_COLUMNS = ("generic_keywords", "generic_keyword")

# ==================== ХЕЛПЕРИ ====================

def load_it_phrases_from_excel(xls: pd.ExcelFile) -> list[str]:
    """
    Чете лист 'it_keywords'.
    Поддържа два формата:
      A) колона 'phrase' (готови фрази, разделени със запетайки)
      B) колони k1..k8 (отделни термина) -> комбинира до 4, разбърква
    Връща списък от вече нормализирани фрази (string), готови за поставяне.
    """
    if "it_keywords" not in xls.sheet_names:
        return []
    df = xls.parse("it_keywords").fillna("")
    df.columns = [str(c).strip().lower() for c in df.columns]
    phrases: list[str] = []
    if "phrase" in df.columns:
        for s in df["phrase"].astype(str):
            s = s.strip().strip(",")
            if s:
                # нормализирай двойни интервали и запетаи
                parts = [p.strip() for p in s.split(",") if p.strip()]
                if parts:
                    import random
                    random.shuffle(parts)
                    phrases.append(", ".join(parts[:4]))
    else:
        # търси k1..k8
        kcols = [c for c in df.columns if re.fullmatch(r"k[1-8]", c)]
        if kcols:
            import random
            for _, row in df[kcols].iterrows():
                terms = [str(row[c]).strip() for c in kcols if str(row[c]).strip()]
                if not terms:
                    continue
                # уникални в оригинален ред
                seen = {}
                terms = [seen.setdefault(t, t) for t in terms if t not in seen]
                random.shuffle(terms)
                phrases.append(", ".join(terms[:4]))
    return phrases

def translate_uk_description_to_it_html(uk_description: str, name: str = "", color: str = "") -> str:
    if not isinstance(uk_description, str):
        return ""
    def norm(s):
        if not isinstance(s, str): return ""
        s = (s.replace("\u2019","'").replace("\u2018","'")
               .replace("\u201c",'"').replace("\u201d",'"'))
        s = s.replace("<br><br>","\n").replace("<br>","\n")
        s = re.sub(r"[ \t]+"," ", s)
        return s.strip()
    text = norm(uk_description)
    if not text: return ""

    # извади име/нюанс ако не са подадени
    def extract_name(src: str) -> str:
        if re.search(r"\[NAME\]", src, flags=re.I):
            return "[NAME]"
        m = re.search(r"for\s+(.+?)\s+-\s+", src, flags=re.I)
        return m.group(1).strip() if m else ""
    def extract_color(src: str) -> str:
        # preserve literal {COLOR} placeholder if present
        if re.search(r"\{COLOR\}", src, flags=re.I):
            return "{COLOR}"
        m = re.search(r"\{([^}]+)\}", src)
        if m: return m.group(1).strip()
        m = re.search(r"\(([^)]+)\)\s+-\s+Rear", src, flags=re.I)
        return m.group(1).strip() if m else ""

    vehicle_name = name.strip() or extract_name(text)
    shade = color.strip() or extract_color(text)

    # ако няма шаблон – върни нормализирания текст като HTML
    key = "enhance your vehicle"
    if key not in text.lower():
        return text.replace("\n","<br>")

    color_it_map = {
        "05% Limo Black": "05% Super Nero",
        "20% Dark Smoke": "20% Fumo Scuro",
        "35% Medium Smoke": "35% Fumo Medio",
        "50% Light Smoke": "50% Fumo Chiaro",
        "70% Ultra Light": "70% Ultra Leggera",
    }
    color_it = color_it_map.get(shade, shade)
    name_disp = vehicle_name or "[NAME]"
    color_frag = f" - ({color_it})" if color_it else ""

    paragraphs = [
        ("<b>TINTCOM Pellicola Oscurante Vetri Auto Pre-Tagliata per "
         f"{name_disp}{color_frag}</b>"),
        ("Migliora comfort, estetica e sicurezza della tua auto con il nostro kit "
         "<b>Pre-Tagliato su Misura</b>. Realizzato per adattarsi perfettamente al modello "
         "della tua vettura, il kit include pellicole per <b>Vetri Posteriori e Lunotto</b>."),
        ("La pellicola blocca fino al 99% dei raggi UV, proteggendo interni, sedili e plancia "
         "da scolorimento e crepe. Riduce l'abbagliamento solare e il calore interno, rendendo "
         "la guida più confortevole durante le giornate calde."),
        ("Il materiale <b>High Performance Series</b> offre visibilità ottimale, mantenendo "
         "l'interno dell'auto più fresco e privato. Grazie alla finitura antigraffio e alla "
         "tecnologia anti-bolle, la pellicola rimane liscia e trasparente nel tempo."),
        ("<b>Facile da installare</b> – non serve smontare pannelli o tagliare la pellicola. "
         "Tutto è pronto per un'installazione rapida fai-da-te con le istruzioni incluse."),
        ("<b>Tonalità disponibili:</b><br>"
         "5% (Super Nero) – massimo oscuramento, massima privacy.<br>"
         "20% (Fumo Scuro) – equilibrio perfetto tra privacy e visibilità.<br>"
         "35% (Fumo Medio) – aspetto elegante e buona visibilità notturna.<br>"
         "50% (Fumo Chiaro) – visibilità eccellente, effetto sobrio e moderno."),
        ("<b>TINTCOM</b> – Soluzioni di protezione auto su misura, con materiali di alta qualità e innovazione.")
    ]
    return "<br><br>".join(paragraphs)

def normalize_html_text(s: str) -> str:
    if not isinstance(s, str):
        return s
    s = (s.replace("\u2019", "'").replace("\u2018", "'")
           .replace("\u201c", '"').replace("\u201d", '"'))
    s = s.replace("<br><br>", "\n").replace("<br>", "\n")
    s = re.sub(r"[ \t]+", " ", s)
    return s.strip()

def load_map_sheet(xls: pd.ExcelFile, sheet_name: str) -> pd.DataFrame:
    if sheet_name not in xls.sheet_names:
        return pd.DataFrame(columns=["field","find","replace"])
    df = xls.parse(sheet_name).fillna("")
    cols = {str(c).strip().lower(): c for c in df.columns}
    df = df.rename(columns={cols.get("field","field"): "field",
                            cols.get("find","find"): "find",
                            cols.get("replace","replace"): "replace"})
    for k in ["field","find","replace"]:
        if k not in df.columns: df[k] = ""
    return df[["field","find","replace"]]

def load_settings(xls: pd.ExcelFile) -> dict:
    # expects sheet "settings" with columns key,value
    res = {}
    if "settings" not in xls.sheet_names:
        return res
    df = xls.parse("settings").fillna("")
    cols = {str(c).strip().lower(): c for c in df.columns}
    key_col = cols.get("key")
    val_col = cols.get("value") or cols.get("val") or cols.get("setting")
    if not key_col:
        return res
    for _, r in df.iterrows():
        k = str(r.get(key_col, "")).strip()
        v = str(r.get(val_col, "")).strip() if val_col else ""
        if k:
            res[k.lower()] = v
    return res

def expand_fields(df_columns, field_cell: str):
    return [c.strip() for c in str(field_cell).split("|")
            if c.strip() and c.strip() in df_columns]

def apply_exact_rules(df, rules: pd.DataFrame):
    for _, r in rules.iterrows():
        cols = expand_fields(df.columns, r.get("field",""))
        find = str(r.get("find",""))
        rep  = str(r.get("replace",""))
        if not cols or find == "":
            continue
        for col in cols:
            s = df[col].astype(str)
            mask = s == find
            cnt = int(mask.sum())
            if cnt:
                df.loc[mask, col] = rep
                APPLIED_LOG[f"{col}::EXACT {find}->{rep}"] += cnt
    return df

def apply_text_rules(df, rules: pd.DataFrame, case_insensitive=True):
    flags = re.IGNORECASE if case_insensitive else 0
    for _, r in rules.iterrows():
        cols = expand_fields(df.columns, r.get("field",""))
        find = str(r.get("find","")).strip()
        rep  = str(r.get("replace",""))
        if not cols or find == "":
            continue
        wordish = re.match(r"^[A-Za-z0-9% .\-]+$", find) is not None
        pattern = r"\b{}\b".format(re.escape(find)) if wordish else re.escape(find)
        for col in cols:
            if col in EXACT_ONLY_FIELDS:
                continue  # никога regex върху системните полета
            if col not in LONG_TEXT_FIELDS:
                continue
            before = df[col].astype(str).map(normalize_html_text)
            after  = before.str.replace(pattern, rep, regex=True, flags=flags)
            changed = int((before != after).sum())
            if changed:
                df[col] = after
                APPLIED_LOG[f"{col}::REGEX {find}->{rep}"] += changed
    return df

def apply_value_map(df, cols, value_map):
    for c in cols:
        if c in df.columns:
            df[c] = df[c].astype(str).map(lambda v: value_map.get(v, v))
    return df

def guess_price_cols(df):
    cols = []
    for c in df.columns:
        lc = str(c).lower()
        if any(h in lc for h in PRICE_COLS_HINTS):
            cols.append(c)
    return cols

def file_out_name(uk_path):
    base = os.path.basename(uk_path)
    if base.lower().startswith("uk-"):
        return os.path.join(os.path.dirname(uk_path), "IT-" + base[3:])
    name, ext = os.path.splitext(base)
    return os.path.join(os.path.dirname(uk_path), f"IT-{name}{ext}")

def cleanup_nulls(df: pd.DataFrame) -> pd.DataFrame:
    return df.map(lambda v: "" if (
        (isinstance(v, float) and pd.isna(v)) or
        (isinstance(v, str) and v.strip().lower() in ("nan","none","nat"))
    ) else v)

def normalize_generic_keyword_columns(
    uk_df: pd.DataFrame,
    it_df: pd.DataFrame,
    default_keywords: str = "",
    heat_shrink_blank: bool = False,
    phrases: list[str] | None = None
) -> pd.DataFrame:
    target_col = next((c for c in GENERIC_KEYWORD_COLUMNS if c in it_df.columns), None)
    if not target_col:
        return it_df

    source_col = next((c for c in GENERIC_KEYWORD_COLUMNS if c in uk_df.columns), None)
    if source_col and source_col != target_col:
        it_df[target_col] = uk_df[source_col].reindex(it_df.index, fill_value="")

    if source_col:
        source_values = uk_df[source_col].reindex(it_df.index, fill_value="")
    else:
        source_values = it_df[target_col]

    source_values = source_values.fillna("").astype(str)
    heat_shrink_mask = source_values.str.strip().str.casefold() == "heat shrink"
    has_uk_kw = source_values.str.strip() != ""

    # по подразбиране – празно навсякъде
    it_df[target_col] = ""

    # ако UK има keywords → поставяме фраза от it_keywords (ако има), иначе fallback към default_keywords
    if phrases:
        ph = phrases[:] if phrases else []
        n = len(ph)
        idx = 0
        sel = it_df[target_col].copy()
        for i in range(len(it_df)):
            if has_uk_kw.iloc[i]:
                if n:
                    sel.iloc[i] = ph[idx % n]
                    idx += 1
                else:
                    sel.iloc[i] = default_keywords
        it_df[target_col] = sel
    else:
        it_df.loc[has_uk_kw, target_col] = default_keywords

    if heat_shrink_blank and heat_shrink_mask.any():
        it_df.loc[heat_shrink_mask, target_col] = ""

    for col in GENERIC_KEYWORD_COLUMNS:
        if col != target_col and col in it_df.columns:
            it_df.drop(columns=col, inplace=True)

    return it_df




def write_into_it_template(template_path: str, df: pd.DataFrame, out_path: str):
    df = df.replace({np.nan: ""}).fillna("")
    wb = load_workbook(template_path)
    ws = wb.active

    # Събери ВСИЧКИ индекси по име (ред 3 са заглавията)
    name_to_cols = {}
    for c in range(1, ws.max_column + 1):
        hdr = ws.cell(row=3, column=c).value
        name_to_cols.setdefault(hdr, []).append(c)

    # Почисти данните (от ред 4 надолу)
    start_row = 4
    if ws.max_row >= start_row:
        for r in range(start_row, ws.max_row + 1):
            for c in range(1, ws.max_column + 1):
                ws.cell(row=r, column=c, value=None)

    # Специален случай: липсва 'part_number' в шаблона, но има 2x 'product_description'
    prod_desc_cols = name_to_cols.get("product_description") or name_to_cols.get("Descrizione prodotto") or []
    has_part_number_col = ("part_number" in name_to_cols)
    two_prod_desc = (len(prod_desc_cols) >= 2)

    r = start_row
    for _, row in df.iterrows():
        # 1) нормално записване: за всяка колона в df пиши във всички целеви колони с това име
        for col_name, val in row.items():
            if col_name in name_to_cols:
                for c in name_to_cols[col_name]:
                    ws.cell(row=r, column=c, value=(None if val == "" else val))

        # 2) специална логика за сбъркан шаблон:
        #    ако нямаме 'part_number' колона в шаблона, но имаме 2х product_description,
        #    прехвърли df['part_number'] във ВТОРАТА product_description колона.
        if (not has_part_number_col) and two_prod_desc and ("part_number" in df.columns):
            pn_val = row.get("part_number", "")
            if isinstance(pn_val, float) and pd.isna(pn_val):
                pn_val = ""
            ws.cell(row=r, column=prod_desc_cols[1], value=(None if pn_val == "" else pn_val))

        r += 1

    wb.save(out_path)


def write_rules_log(out_path):
    out_dir = os.path.dirname(out_path) or "."
    path = os.path.join(out_dir, "applied_rules_log.csv")
    with open(path, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f); w.writerow(["field","rule","count"])
        for k, v in APPLIED_LOG.items():
            field, rule = (k.split("::",1)+[""])[:2]
            w.writerow([field, rule, v])
    return path

# ==================== ОСНОВНА ФУНКЦИЯ ====================

def uk_to_it(uk_file, it_template_file, out_path=None, find_replace_xlsx=None):
    global ITEM_TAIL, GENERIC_KEYWORDS_DEFAULT
    # 1) UK таблица (ред 3 = header)
    uk = pd.read_excel(uk_file, header=2, dtype=str).fillna("")

    # 2) IT шаблон (взимаме имената от ред 3)
    it_raw = pd.read_excel(it_template_file, header=None, dtype=str).fillna("")
    it_header_names = list(it_raw.iloc[2])
    it_out = pd.DataFrame(columns=it_header_names)

    # 3) Копиране по име
    cols = COLUMNS_TO_COPY or [c for c in uk.columns if c in it_out.columns]
    for c in cols:
        it_out[c] = uk[c]

    # default empty tone map
    tone_map = {}
    # default settings
    settings = {}
    generic_kw_default = GENERIC_KEYWORDS_DEFAULT
    heat_shrink_blank_flag = False
    item_tail_from_xls = ""
    # 4) Правила
    phrases_for_it = []
    if find_replace_xlsx:
        xls = pd.ExcelFile(find_replace_xlsx)
        # settings & tone map
        settings = load_settings(xls)
        item_tail_from_xls = settings.get("item_tail_it", "").strip() or item_tail_from_xls
        generic_kw_default = settings.get("generic_keywords_default", "").strip() or generic_kw_default
        heat_shrink_blank_flag = (settings.get("heat_shrink_blank", "0").strip().lower() in ("1","true","yes","y"))
        # зареди италианските ключови фрази (ако има лист)
        phrases_for_it = load_it_phrases_from_excel(xls)

        text_map  = load_map_sheet(xls, "text_map")
        words_map = load_map_sheet(xls, "words_find_replace") if "words_find_replace" in xls.sheet_names else (
                    load_map_sheet(xls, "words_find_replece") if "words_find_replece" in xls.sheet_names
                    else pd.DataFrame(columns=["field","find","replace"]))
        sku_map   = load_map_sheet(xls, "sku_map")
        price_map_df = xls.parse("price_map").fillna("") if "price_map" in xls.sheet_names else pd.DataFrame(columns=["find","replace"])

        # tone_map sheet (find -> replace) applied to color_name/size_name
        if "tone_map" in xls.sheet_names:
            tm = xls.parse("tone_map").fillna("")
            # find/replace columns tolerant lookup
            cols_map = {str(c).strip().lower(): c for c in tm.columns}
            find_col = cols_map.get("find") or cols_map.get("from") or list(tm.columns)[0]
            rep_col  = cols_map.get("replace") or cols_map.get("to") or (list(tm.columns)[1] if len(tm.columns)>1 else find_col)
            tone_map = dict(zip(tm[find_col].astype(str), tm[rep_col].astype(str)))
            # apply immediately to color_name / size_name (if present)
            it_out = apply_value_map(it_out, [c for c in ("color_name","size_name") if c in it_out.columns], tone_map)

        # 4.1 точни замени
        it_out = apply_exact_rules(it_out, text_map)
        it_out = apply_exact_rules(it_out, words_map)

        # 4.2 sku точни замени
        for _, r in sku_map.iterrows():
            cols_ = expand_fields(it_out.columns, r.get("field",""))
            find = str(r.get("find","")); rep = str(r.get("replace",""))
            if not cols_ or find == "": continue
            for col in cols_:
                s = it_out[col].astype(str)
                mask = s == find
                cnt = int(mask.sum())
                if cnt:
                    it_out.loc[mask, col] = rep
                    APPLIED_LOG[f"{col}::SKU_MAP {find}->{rep}"] += cnt

        # 4.3 regex – само за дълги текстови полета
        it_out = apply_text_rules(it_out, text_map)
        it_out = apply_text_rules(it_out, words_map)

        # 4.4 цени
        price_cols = guess_price_cols(it_out)
        value_map  = dict(zip(price_map_df["find"].astype(str), price_map_df["replace"].astype(str)))
        it_out     = apply_value_map(it_out, price_cols, value_map)

    # 5) Бизнес правила
    if heat_shrink_blank_flag and "installation_type" in it_out.columns:
        mask = it_out["installation_type"].astype(str).str.strip().str.lower().eq("heat shrink")
        it_out.loc[mask, "installation_type"] = ""
    if "item_name" in it_out.columns:
        it_out["item_name"] = it_out["item_name"].astype(str).map(
        lambda s: normalize_item_name(s, item_tail_from_xls)
    )

    # NOTE: material_type and color_map hard-coded replacements moved to Find_and_Replace.xlsx (text_map / words_find_replace)
    # remove in-code replacements; if users want them they must place appropriate rules in the Excel.

    # apply tone_map again in case it was provided but columns added later
    if 'tone_map' in locals() and tone_map:
        it_out = apply_value_map(it_out, [c for c in ("color_name","size_name") if c in it_out.columns], tone_map)

    # „Heat Shrink“ -> празно controlled by settings (heat_shrink_blank = "1"/"0")
    heat_shrink_flag = True
    if isinstance(settings, dict) and settings.get("heat_shrink_blank") is not None:
        v = str(settings.get("heat_shrink_blank","")).strip().lower()
        heat_shrink_flag = v in ("1","true","yes","y")
    if heat_shrink_flag and "installation_type" in it_out.columns:
        mask = it_out["installation_type"].astype(str).str.strip().str.lower().eq("heat shrink")
        it_out.loc[mask, "installation_type"] = ""

    # product_description: генерирай италиански HTML по UK шаблона (ако има)
    if "product_description" in it_out.columns:
        it_out["product_description"] = it_out.apply(
            lambda row: translate_uk_description_to_it_html(
                row.get("product_description",""),
                row.get("item_name",""),
                row.get("color_name","")
            ),
            axis=1
        )

    # keywords – попълни правилната колона:
    # само за редове, където UK има keywords; използвай it_keywords (разбъркани до 4 термина)
    it_out = normalize_generic_keyword_columns(
        uk, it_out,
        default_keywords=generic_kw_default,
        heat_shrink_blank=heat_shrink_blank_flag,
        phrases=phrases_for_it
    )

    # 6) Почистване и запис
    it_out = cleanup_nulls(it_out)
    out_file = out_path or file_out_name(uk_file)
    write_into_it_template(it_template_file, it_out, out_file)
    log_path = write_rules_log(out_file)
    print(f"Done → {out_file}")
    print(f"Rules log → {log_path}")

# ==================== RUN ====================
if __name__ == "__main__":
    if len(sys.argv) < 3:
        print("Usage: python uk_to_it.py <UK.xlsx> <IT_template.xlsx> [Find_and_Replace.xlsx] [OUT.xlsx]")
        sys.exit(1)
    uk   = sys.argv[1]
    it_t = sys.argv[2]
    fr   = sys.argv[3] if len(sys.argv) >= 4 else None
    out  = sys.argv[4] if len(sys.argv) >= 5 else None
    uk_to_it(uk, it_t, out, fr)